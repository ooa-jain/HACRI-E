"""
The single-tab department summary: Registered, Baseline, Post, Deeksharambh —
one row per department, one workbook. Added next to the roster export on the
overview page so the whole cohort can be scanned or sorted without opening a
report per department.
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import AsyncIterator

import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from mongomock_motor import AsyncMongoMockClient
from openpyxl import load_workbook

from app import db

LAW = "Department of Law"
COMMERCE = "Department of Commerce"


@pytest_asyncio.fixture
async def client() -> AsyncIterator[AsyncClient]:
    db._set_client_for_tests(AsyncMongoMockClient())
    try:
        from app.main import app
        await db.init_indexes(allow_duplicate_email=True)
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as ac:
            yield ac
    finally:
        db._reset_clients_for_tests()


async def _seed() -> None:
    now = datetime.now(timezone.utc)

    async def user(email, dept, status):
        await db.get_db()["users"].insert_one({
            "email": email, "name": email, "program": dept, "ug_or_pg": "ug",
            "status": status, "created_at": now})

    # Law: 3 registered, 2 baseline, 1 post, 1 orientation (a different student
    # from the one who finished post — orientation is its own flow).
    await user("law1@x.com", LAW, db.STATUS_PRE_DONE)
    await user("law2@x.com", LAW, db.STATUS_POST_DONE)
    await user("law3@x.com", LAW, None)
    await db.get_db()["orientation_responses"].insert_one(
        {"email": "law1@x.com", "name": "law1", "submitted_at": now, "data": {}})

    # Commerce: 1 registered, 1 baseline, 0 post, 0 orientation.
    await user("com1@x.com", COMMERCE, db.STATUS_PRE_DONE)

    # A reply from nobody we have a registration record for — filed under
    # "Other" rather than dropped.
    await db.get_db()["orientation_responses"].insert_one(
        {"email": "ghost@x.com", "name": "Ghost", "submitted_at": now, "data": {}})


# ── The data function ─────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_each_department_gets_all_four_counts(client):
    await _seed()
    summary = await db.department_registration_summary()
    rows = {r["dept"]: r for r in summary["departments"]}

    law = rows[LAW]
    assert law["registered"] == 3
    assert law["pre_done"] == 2 and law["pre_pending"] == 1
    assert law["post_done"] == 1 and law["post_pending"] == 1
    assert law["orientation_done"] == 1 and law["orientation_pending"] == 2

    com = rows[COMMERCE]
    assert com["registered"] == 1
    assert com["pre_done"] == 1
    assert com["orientation_done"] == 0 and com["orientation_pending"] == 1


@pytest.mark.asyncio
async def test_an_orientation_reply_with_no_registration_record_is_not_dropped(client):
    await _seed()
    summary = await db.department_registration_summary()
    other = next(r for r in summary["departments"] if r["dept"] == "Other")
    assert other["orientation_done"] == 1
    assert other["registered"] == 0   # the ghost reply, nothing else


@pytest.mark.asyncio
async def test_a_resubmitted_orientation_form_counts_the_student_once(client):
    await _seed()
    now = datetime.now(timezone.utc)
    await db.get_db()["orientation_responses"].insert_one(
        {"email": "law1@x.com", "name": "law1", "submitted_at": now, "data": {"q2": 9}})

    summary = await db.department_registration_summary()
    law = next(r for r in summary["departments"] if r["dept"] == LAW)
    assert law["orientation_done"] == 1


@pytest.mark.asyncio
async def test_totals_sum_every_department(client):
    await _seed()
    summary = await db.department_registration_summary()
    assert summary["totals"]["registered"] == sum(
        r["registered"] for r in summary["departments"])
    assert summary["totals"]["orientation_done"] == 2   # law1 + ghost


# ── The route ────────────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_the_export_needs_an_admin_session(client):
    r = await client.get("/admin/survey/export-department-summary")
    assert r.status_code == 403


@pytest.mark.asyncio
async def test_the_workbook_has_one_sheet_with_every_department(client):
    await _seed()
    client.cookies.set("survey_admin_session", "1")

    r = await client.get("/admin/survey/export-department-summary")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert ".xlsx" in r.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["Department Summary"]
    ws = wb.active

    header = [c.value for c in ws[4]]
    assert header == ["Department", "Registered", "Baseline done", "Baseline pending",
                      "Post survey done", "Post survey pending",
                      "Deeksharambh done", "Deeksharambh pending"]

    body = [[c.value for c in row] for row in ws.iter_rows(min_row=5)]
    depts = [row[0] for row in body]
    assert "All departments" in depts   # the cohort total leads the sheet
    assert LAW in depts and COMMERCE in depts

    law_row = next(row for row in body if row[0] == LAW)
    assert law_row[1:] == [3, 2, 1, 1, 1, 1, 2]
