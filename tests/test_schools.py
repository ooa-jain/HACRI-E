"""
Schools — the departments folded up one level.

A department is what a student picks at registration; a school is the unit a
dean asks about. Until this existed, answering "how is the School of Sciences
doing?" meant adding up seven departments by hand.
"""

from __future__ import annotations

from datetime import datetime, timedelta, timezone
from typing import AsyncIterator

import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from mongomock_motor import AsyncMongoMockClient

from app import db
from app.schools import OTHER_SCHOOL, SCHOOLS, school_of, school_slug

SCIENCES = "School of Sciences"
LAW = "School of Law"


@pytest_asyncio.fixture
async def app_with_mock():
    db._set_client_for_tests(AsyncMongoMockClient())
    try:
        from app.main import app
        await db.init_indexes(allow_duplicate_email=True)
        yield app
    finally:
        db._reset_clients_for_tests()


@pytest_asyncio.fixture
async def client(app_with_mock) -> AsyncIterator[AsyncClient]:
    transport = ASGITransport(app=app_with_mock)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        yield ac


@pytest_asyncio.fixture
async def admin(client) -> AsyncClient:
    client.cookies.set("survey_admin_session", "1")
    return client


async def _student(email: str, *, program: str, status: str | None = None,
                   days_ago: float = 1, score: int = 3) -> None:
    from app.hacri_e2_compat import SCHEMA

    when = datetime.now(timezone.utc) - timedelta(days=days_ago)
    await db.get_db()["users"].insert_one({
        "email": email, "name": email.split("@")[0], "program": program,
        "ug_or_pg": "ug", "status": status, "created_at": when,
        "pre_submitted_at": when if status else None,
    })
    if status:
        await db.get_db()["pre_responses"].insert_one({
            "email": email, "name": email, "submitted_at": when,
            "fields": {k: score for k in SCHEMA}})
    if status == db.STATUS_POST_DONE:
        await db.get_db()["post_responses"].insert_one({
            "email": email, "name": email, "submitted_at": when,
            "fields": {k: score for k in SCHEMA}})


# ── The mapping ──────────────────────────────────────────────────────────────

def test_every_registered_department_belongs_to_a_school():
    """A department in "Other" is a real signal, not a bug — but only CeRSSE
    should be there, because it appears in no school's list."""
    from app.departments import DEPARTMENTS

    homeless = [d for d in DEPARTMENTS if school_of(d) == OTHER_SCHOOL]
    assert homeless == ["CeRSSE"], homeless


def test_a_department_is_claimed_by_exactly_one_school():
    seen: dict[str, str] = {}
    for school, departments in SCHOOLS.items():
        for dept in departments:
            assert dept not in seen, f"{dept} is in both {seen.get(dept)} and {school}"
            seen[dept] = school


def test_the_mapping_survives_spelling_differences():
    """& vs and, case and punctuation must not lose a department."""
    assert school_of("Department of Law") == LAW
    assert school_of("department of law") == LAW
    assert school_of("Department of Humanities and Social Sciences") == \
        "School of Humanities and Social Sciences"
    assert school_of("Department of Humanities & Social Sciences") == \
        "School of Humanities and Social Sciences"


def test_anything_unrecognised_lands_in_other_rather_than_vanishing():
    assert school_of("Department of Something Nobody Listed") == OTHER_SCHOOL
    assert school_of("") == OTHER_SCHOOL
    assert school_of(None) == OTHER_SCHOOL


def test_school_slugs_are_url_safe():
    assert school_slug("School of Engineering & Technology") == \
        "school-of-engineering-technology"
    assert school_slug("CMS Business School") == "cms-business-school"


# ── The figures ──────────────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_a_school_total_is_its_departments_added_up(app_with_mock):
    await _student("a@x.com", program="Department of Forensic Science",
                   status=db.STATUS_PRE_DONE)
    await _student("b@x.com", program="Department of Physics and Electronics",
                   status=db.STATUS_POST_DONE)
    await _student("c@x.com", program="Department of Forensic Science")  # not started
    await _student("d@x.com", program="Department of Law", status=db.STATUS_PRE_DONE)

    data = await db.get_school_analysis_data()
    rows = {r["school"]: r for r in data["schools"]}

    sciences = rows[SCIENCES]
    assert sciences["registered"] == 3
    assert sciences["pre_done"] == 2          # pre_done + post_done students
    assert sciences["post_done"] == 1
    assert sciences["dept_count"] == 2        # only departments with students
    # The fold is exact: the school equals the sum of its own departments.
    assert sciences["registered"] == sum(d["registered"] for d in sciences["departments"])
    assert sciences["pre_done"] == sum(d["pre_done"] for d in sciences["departments"])

    assert rows[LAW]["registered"] == 1


@pytest.mark.asyncio
async def test_an_unlisted_department_is_counted_under_other(app_with_mock):
    await _student("x@x.com", program="CeRSSE", status=db.STATUS_PRE_DONE)
    await _student("y@x.com", program="Some Department We Never Heard Of")

    rows = {r["school"]: r for r in (await db.get_school_analysis_data())["schools"]}
    assert rows[OTHER_SCHOOL]["registered"] == 2
    assert rows[OTHER_SCHOOL]["pre_done"] == 1


@pytest.mark.asyncio
async def test_averages_are_weighted_by_students_not_by_department(app_with_mock):
    """A department of one must not weigh the same as a department of ten."""
    for i in range(10):
        await _student(f"big{i}@x.com", program="Department of Forensic Science",
                       status=db.STATUS_PRE_DONE, score=2)
    await _student("small@x.com", program="Department of Physics and Electronics",
                   status=db.STATUS_PRE_DONE, score=5)

    rows = {r["school"]: r for r in (await db.get_school_analysis_data())["schools"]}
    average = rows[SCIENCES]["avg_lit_pre"]

    # A plain mean of the two department averages would sit halfway; weighting
    # by students pulls it towards the ten.
    depts = {d["dept"]: d for d in rows[SCIENCES]["departments"]}
    low = depts["Department of Forensic Science"]["avg_lit_pre"]
    high = depts["Department of Physics and Electronics"]["avg_lit_pre"]
    assert low < average < high
    assert abs(average - low) < abs(average - high)


@pytest.mark.asyncio
async def test_recent_submissions_are_counted_separately_from_the_total(app_with_mock):
    """A school that finished in June and one filling today must not look alike."""
    await _student("old@x.com", program="Department of Law",
                   status=db.STATUS_PRE_DONE, days_ago=60)
    await _student("new@x.com", program="Department of Forensic Science",
                   status=db.STATUS_PRE_DONE, days_ago=1)

    data = await db.get_school_analysis_data()
    rows = {r["school"]: r for r in data["schools"]}

    assert rows[LAW]["pre_done"] == 1 and rows[LAW]["recent_total"] == 0
    assert rows[SCIENCES]["recent_total"] == 1
    assert rows[SCIENCES]["last_submission"]
    assert data["highlights"]["most_recent"]["school"] == SCIENCES


@pytest.mark.asyncio
async def test_a_school_with_nobody_registered_is_still_listed(app_with_mock):
    await _student("only@x.com", program="Department of Law", status=db.STATUS_PRE_DONE)

    data = await db.get_school_analysis_data()
    names = [r["school"] for r in data["schools"]]

    assert "School of Aviation and Aerospace Management" in names
    assert data["overall"]["school_count"] == 1        # only one has students
    assert data["overall"]["schools_listed"] == len(SCHOOLS) + 1   # +Other


@pytest.mark.asyncio
async def test_the_highlights_name_the_busiest_and_the_quietest(app_with_mock):
    for i in range(5):
        await _student(f"law{i}@x.com", program="Department of Law",
                       status=db.STATUS_POST_DONE)
    await _student("quiet@x.com", program="Department of Commerce",
                   status=db.STATUS_PRE_DONE)
    for i in range(4):
        await _student(f"stalled{i}@x.com", program="Department of Forensic Science")

    hi = (await db.get_school_analysis_data())["highlights"]
    assert hi["most_submissions"]["school"] == LAW               # 10 submissions
    # Four students registered and not one has started: fewest submissions and
    # the worst completion rate are the same school here, and both are the
    # answer to "who needs chasing".
    assert hi["fewest_submissions"]["school"] == SCIENCES
    assert hi["fewest_submissions"]["registered"] == 4
    assert hi["lowest_completion"]["school"] == SCIENCES
    assert hi["highest_completion"]["school"] in (LAW, "School of Commerce")


# ── The admin page ───────────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_the_school_api_needs_an_admin_session(client: AsyncClient):
    assert (await client.get("/admin/api/survey/school-analysis")).status_code == 403


@pytest.mark.asyncio
async def test_the_admin_page_offers_a_link_per_school_and_one_for_all(admin: AsyncClient):
    await _student("a@x.com", program="Department of Law", status=db.STATUS_PRE_DONE)

    body = (await admin.get("/admin/api/survey/school-analysis")).json()

    assert body["overall"]["directory_url"].endswith(
        body["overall"]["directory_token"])
    assert "/shared/schools?token=" in body["overall"]["directory_url"]

    law = next(r for r in body["schools"] if r["school"] == LAW)
    # One link per school, covering both surveys, plus its own workbook.
    assert "/shared/school?school=" in law["share_url"]
    assert "/shared/school/export-excel?school=" in law["excel_url"]
    assert law["token"] in law["share_url"] and law["token"] in law["excel_url"]


# ── The shared pages ─────────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_the_all_schools_page_needs_its_token(client: AsyncClient):
    from app.routes.shared_analysis import get_schools_directory_token

    assert (await client.get("/shared/schools?token=nope")).status_code == 403

    await _student("a@x.com", program="Department of Law", status=db.STATUS_PRE_DONE)
    ok = await client.get("/shared/schools",
                          params={"token": get_schools_directory_token()})
    assert ok.status_code == 200
    assert LAW in ok.text
    assert "Share of all submissions" in ok.text       # the pie
    assert "Submissions by school" in ok.text          # and the ranked bars


@pytest.mark.asyncio
async def test_one_schools_link_cannot_be_edited_into_anothers(client: AsyncClient):
    from app.routes.shared_analysis import get_school_token

    await _student("a@x.com", program="Department of Law", status=db.STATUS_PRE_DONE)

    law_token = get_school_token(LAW)
    assert (await client.get("/shared/school",
                             params={"school": LAW, "token": law_token})).status_code == 200
    # The same token pointed at another school opens nothing.
    assert (await client.get("/shared/school",
                             params={"school": SCIENCES, "token": law_token})).status_code == 403
    assert (await client.get("/shared/school",
                             params={"school": LAW, "token": "guessed"})).status_code == 403
    # Links handed out before the two reports were merged still open the
    # merged one, rather than becoming dead mail.
    for old in ("pre", "post"):
        assert (await client.get(
            "/shared/school",
            params={"school": LAW, "token": get_school_token(LAW, old)})).status_code == 200


@pytest.mark.asyncio
async def test_a_school_page_lists_the_departments_inside_it(client: AsyncClient):
    from app.routes.shared_analysis import get_school_token

    await _student("a@x.com", program="Department of Forensic Science",
                   status=db.STATUS_PRE_DONE)
    await _student("b@x.com", program="Department of Physics and Electronics",
                   status=db.STATUS_PRE_DONE)

    page = await client.get("/shared/school",
                            params={"school": SCIENCES,
                                    "token": get_school_token(SCIENCES)})
    assert page.status_code == 200
    assert "Department of Forensic Science" in page.text
    assert "Department of Physics and Electronics" in page.text
    # And no student is named on a page anyone with the link can open.
    assert "a@x.com" not in page.text


@pytest.mark.asyncio
async def test_an_unknown_school_is_not_found(client: AsyncClient):
    from app.routes.shared_analysis import get_school_token

    resp = await client.get("/shared/school",
                            params={"school": "School of Atlantis",
                                    "token": get_school_token("School of Atlantis")})
    assert resp.status_code == 404


# ── One report, both surveys ─────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_the_school_report_shows_baseline_and_post_together(client: AsyncClient):
    """Reading the change used to mean opening two tabs and subtracting by eye."""
    from app.routes.shared_analysis import get_school_token

    await _student("low@x.com", program="Department of Law",
                   status=db.STATUS_POST_DONE, score=2)
    # Nudge the post answers up, so there is a change to state.
    from app.hacri_e2_compat import SCHEMA
    await db.get_db()["post_responses"].update_one(
        {"email": "low@x.com"}, {"$set": {"fields": {k: 4 for k in SCHEMA}}})

    page = await client.get("/shared/school",
                            params={"school": LAW, "token": get_school_token(LAW)})
    assert page.status_code == 200

    text = page.text
    assert "Baseline" in text and "Post-workshop" in text
    assert "What changed" in text                # the section it lives under
    assert "after the workshop" in text          # the change, in words
    assert "Export Excel" in text


@pytest.mark.asyncio
async def test_the_change_is_not_invented_when_only_one_survey_is_in(client: AsyncClient):
    from app.routes.shared_analysis import get_school_token

    await _student("only@x.com", program="Department of Law",
                   status=db.STATUS_PRE_DONE)

    page = await client.get("/shared/school",
                            params={"school": LAW, "token": get_school_token(LAW)})
    assert "Waiting on both surveys" in page.text
    assert "after the workshop" not in page.text


# ── The workbook ─────────────────────────────────────────────────────────────

def _sheets(book: bytes) -> list[str]:
    import io as _io
    from openpyxl import load_workbook
    return load_workbook(_io.BytesIO(book)).sheetnames


@pytest.mark.asyncio
async def test_the_workbook_opens_all_schools_then_departments_then_a_tab_each(app_with_mock):
    from app.school_export import generate_schools_excel

    await _student("a@x.com", program="Department of Law", status=db.STATUS_PRE_DONE)
    await _student("b@x.com", program="Department of Forensic Science",
                   status=db.STATUS_POST_DONE)
    await _student("c@x.com", program="Department of Physics and Electronics")

    data = await db.get_school_analysis_data()
    names = _sheets(generate_schools_excel(data, generated_at="today"))

    assert names[0] == "All Schools"
    assert names[1] == "Departments"
    assert LAW in names and SCIENCES in names
    # A school nobody registered under has nothing to put on a tab.
    assert "School of Aviation and Aerospace Management" not in names


@pytest.mark.asyncio
async def test_every_department_appears_on_the_departments_sheet(app_with_mock):
    import io as _io
    from openpyxl import load_workbook
    from app.school_export import generate_schools_excel

    await _student("a@x.com", program="Department of Law", status=db.STATUS_PRE_DONE)
    await _student("b@x.com", program="Department of Forensic Science",
                   status=db.STATUS_PRE_DONE)
    await _student("c@x.com", program="CeRSSE")

    data = await db.get_school_analysis_data()
    wb = load_workbook(_io.BytesIO(generate_schools_excel(data, generated_at="t")))
    ws = wb["Departments"]

    listed = {row[1] for row in ws.iter_rows(min_row=5, values_only=True) if row[1]}
    assert "Department of Law" in listed
    assert "Department of Forensic Science" in listed
    assert "CeRSSE" in listed            # the unmapped one is not dropped

    # Baseline and post sit side by side with the change between them.
    headers = [c.value for c in ws[4]]
    assert "Avg literacy (baseline)" in headers
    assert "Avg literacy (post)" in headers
    assert "Literacy change" in headers


@pytest.mark.asyncio
async def test_a_school_tab_totals_its_own_departments(app_with_mock):
    import io as _io
    from openpyxl import load_workbook
    from app.school_export import generate_schools_excel

    for i in range(3):
        await _student(f"f{i}@x.com", program="Department of Forensic Science",
                       status=db.STATUS_PRE_DONE)
    await _student("p@x.com", program="Department of Physics and Electronics",
                   status=db.STATUS_POST_DONE)

    data = await db.get_school_analysis_data()
    wb = load_workbook(_io.BytesIO(generate_schools_excel(data, generated_at="t")))
    ws = wb[SCIENCES]

    rows = [r for r in ws.iter_rows(min_row=5, values_only=True) if r[0]]
    total = next(r for r in rows if str(r[0]).startswith("TOTAL"))
    departments = [r for r in rows if not str(r[0]).startswith("TOTAL")]

    assert len(departments) == 2
    assert total[1] == sum(d[1] for d in departments)     # registered
    assert total[2] == sum(d[2] for d in departments)     # baseline done


@pytest.mark.asyncio
async def test_the_workbooks_download_from_the_admin_and_the_shared_links(admin: AsyncClient):
    from app.routes.shared_analysis import get_school_token, get_schools_directory_token

    await _student("a@x.com", program="Department of Law", status=db.STATUS_PRE_DONE)

    XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    from_admin = await admin.get("/admin/survey/export-schools")
    assert from_admin.status_code == 200
    assert from_admin.headers["content-type"] == XLSX
    assert "All Schools" in _sheets(from_admin.content)

    shared = await admin.get("/shared/schools/export-excel",
                             params={"token": get_schools_directory_token()})
    assert shared.status_code == 200
    assert _sheets(shared.content)[0] == "All Schools"

    one = await admin.get("/shared/school/export-excel",
                          params={"school": LAW, "token": get_school_token(LAW)})
    assert one.status_code == 200
    assert _sheets(one.content) == [LAW]


@pytest.mark.asyncio
async def test_the_exports_are_not_open_to_anyone_without_the_link(client: AsyncClient):
    assert (await client.get("/admin/survey/export-schools")).status_code == 403
    assert (await client.get("/shared/schools/export-excel",
                             params={"token": "guessed"})).status_code == 403
    assert (await client.get("/shared/school/export-excel",
                             params={"school": LAW, "token": "guessed"})).status_code == 403
