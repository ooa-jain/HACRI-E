"""
The "All data" export: like the Full report's department tabs, but carrying
the actual Pre/Post/Deeksharambh answers behind each student's done/pending
status — question by question, plus the literacy/readiness scores those
answers work out to.
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import AsyncIterator

import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from mongomock_motor import AsyncMongoMockClient
from openpyxl import load_workbook

from app import db
from app.hacri_e2_compat import SCHEMA

LAW = "Department of Law"
COMMERCE = "Department of Commerce"
LIT_KEY = next(k for k, (kind, _) in SCHEMA.items() if kind == "L")


@pytest_asyncio.fixture
async def client() -> AsyncIterator[AsyncClient]:
    db._set_client_for_tests(AsyncMongoMockClient())
    try:
        from app.main import app
        await db.init_indexes(allow_duplicate_email=True)
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as ac:
            yield ac
    finally:
        db._reset_clients_for_tests()


async def _seed() -> None:
    now = datetime.now(timezone.utc)

    async def user(email, dept, status):
        await db.get_db()["users"].insert_one({
            "email": email, "name": email.split("@")[0], "program": dept,
            "ug_or_pg": "ug", "status": status, "created_at": now})

    await user("law1@x.com", LAW, db.STATUS_POST_DONE)
    await db.get_db()["pre_responses"].insert_one(
        {"email": "law1@x.com", "submitted_at": now, "fields": {LIT_KEY: "4", "A1": "18-20"}})
    await db.get_db()["post_responses"].insert_one(
        {"email": "law1@x.com", "submitted_at": now, "fields": {LIT_KEY: "5", "father_name": "Mr. Law"}})
    await db.get_db()["orientation_responses"].insert_one({
        "email": "law1@x.com", "name": "law1", "submitted_at": now,
        "data": {"q1": ["Fun", "Fast-paced"], "q2": 9},
    })

    await user("com1@x.com", COMMERCE, None)


@pytest.mark.asyncio
async def test_master_report_carries_the_raw_answers(client):
    await _seed()
    report = await db.department_master_report()

    law = next(d for d in report["departments"] if d["dept"] == LAW)
    law1 = next(s for s in law["students"] if s["email"] == "law1@x.com")
    assert law1["pre_fields"][LIT_KEY] == "4"
    assert law1["post_fields"][LIT_KEY] == "5"
    assert law1["orientation_data"]["q2"] == 9

    assert LIT_KEY in report["schema_keys"]
    assert "A1" in report["pre_keys"]          # non-Likert key picked up too
    assert "father_name" in report["post_keys"]
    assert report["orientation_labels"]["q2"] == "Overall vibe of Deeksharambh 2026"


@pytest.mark.asyncio
async def test_the_export_needs_an_admin_session(client):
    r = await client.get("/admin/survey/export-all-data")
    assert r.status_code == 403


@pytest.mark.asyncio
async def test_the_workbook_has_question_level_columns_per_department(client):
    await _seed()
    client.cookies.set("survey_admin_session", "1")

    r = await client.get("/admin/survey/export-all-data")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]

    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames[0] == "Overview"
    assert wb.sheetnames[1] == "Charts"
    assert LAW in wb.sheetnames

    law_ws = wb[LAW]
    header = [c.value for c in law_ws[4]]
    assert header[:4] == ["Name", "Email", "Level", "Registered"]
    assert "PRE Literacy" in header
    assert f"PRE {LIT_KEY}" in header
    assert f"POST {LIT_KEY}" in header
    assert f"Δ {LIT_KEY}" in header
    assert "Deeksharambh: Overall vibe of Deeksharambh 2026" in header

    body_row = next(
        [c.value for c in r_] for r_ in law_ws.iter_rows(min_row=5)
        if r_[1].value == "law1@x.com"
    )
    row_by_header = dict(zip(header, body_row))
    assert row_by_header[f"PRE {LIT_KEY}"] == "4"
    assert row_by_header[f"POST {LIT_KEY}"] == "5"
    assert row_by_header[f"Δ {LIT_KEY}"] == 1
    assert row_by_header["Deeksharambh: Overall vibe of Deeksharambh 2026"] == 9
    assert row_by_header["Deeksharambh: Words describing the Deeksharambh experience"] == "Fun; Fast-paced"
