"""
The "whole report in one Excel" export: an Overview tab (same numbers as the
single-tab department summary), a Charts tab (pie + bar), and one tab per
department — descending by registered count — carrying that department's
summary plus every student behind it.
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import AsyncIterator

import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from mongomock_motor import AsyncMongoMockClient
from openpyxl import load_workbook

from app import db
from app.excel_export import _safe_sheet_name

LAW = "Department of Law"
COMMERCE = "Department of Commerce"


@pytest_asyncio.fixture
async def client() -> AsyncIterator[AsyncClient]:
    db._set_client_for_tests(AsyncMongoMockClient())
    try:
        from app.main import app
        await db.init_indexes(allow_duplicate_email=True)
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as ac:
            yield ac
    finally:
        db._reset_clients_for_tests()


async def _seed() -> None:
    now = datetime.now(timezone.utc)

    async def user(email, dept, status):
        await db.get_db()["users"].insert_one({
            "email": email, "name": email.split("@")[0], "program": dept,
            "ug_or_pg": "ug", "status": status, "created_at": now})

    # Law: 3 registered, 2 baseline, 1 post, 1 orientation.
    await user("law1@x.com", LAW, db.STATUS_PRE_DONE)
    await user("law2@x.com", LAW, db.STATUS_POST_DONE)
    await user("law3@x.com", LAW, None)
    await db.get_db()["orientation_responses"].insert_one(
        {"email": "law1@x.com", "name": "law1", "submitted_at": now, "data": {}})

    # Commerce: 1 registered, 1 baseline.
    await user("com1@x.com", COMMERCE, db.STATUS_PRE_DONE)

    # A reply from nobody we have a registration record for.
    await db.get_db()["orientation_responses"].insert_one(
        {"email": "ghost@x.com", "name": "Ghost", "submitted_at": now, "data": {}})


# ── The data function ─────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_department_full_report_matches_the_summary_totals(client):
    await _seed()
    report = await db.department_full_report()
    summary_depts = {r["dept"] for r in report["summary"]["departments"]}
    report_depts = {d["dept"] for d in report["departments"]}
    assert summary_depts == report_depts

    law = next(d for d in report["departments"] if d["dept"] == LAW)
    assert len(law["students"]) == 3
    emails = {s["email"] for s in law["students"]}
    assert emails == {"law1@x.com", "law2@x.com", "law3@x.com"}

    law1 = next(s for s in law["students"] if s["email"] == "law1@x.com")
    assert law1["pre_done"] is True
    assert law1["orientation_done"] is True
    law3 = next(s for s in law["students"] if s["email"] == "law3@x.com")
    assert law3["pre_done"] is False and law3["orientation_done"] is False


@pytest.mark.asyncio
async def test_departments_are_ordered_the_same_as_the_summary(client):
    await _seed()
    report = await db.department_full_report()
    summary_order = [r["dept"] for r in report["summary"]["departments"]]
    report_order = [d["dept"] for d in report["departments"]]
    assert summary_order == report_order


@pytest.mark.asyncio
async def test_a_ghost_orientation_reply_lands_on_the_other_tab(client):
    await _seed()
    report = await db.department_full_report()
    other = next(d for d in report["departments"] if d["dept"] == "Other")
    ghost = next(s for s in other["students"] if s["email"] == "ghost@x.com")
    assert ghost["orientation_done"] is True
    assert ghost["pre_done"] is False


# ── Sheet-name safety ────────────────────────────────────────────────────

def test_safe_sheet_name_strips_forbidden_characters():
    used: set[str] = set()
    name = _safe_sheet_name("A/B*C?D:E[F]G", used)
    assert not any(ch in name for ch in "\\/*?:[]")


def test_safe_sheet_name_dedupes_long_similar_names():
    used: set[str] = set()
    a = _safe_sheet_name("Department of Management Studies - UG", used)
    b = _safe_sheet_name("Department of Management Studies - PG", used)
    assert a != b
    assert len(a) <= 31 and len(b) <= 31


# ── The route ────────────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_the_export_needs_an_admin_session(client):
    r = await client.get("/admin/survey/export-full-report")
    assert r.status_code == 403


@pytest.mark.asyncio
async def test_the_workbook_has_overview_charts_and_a_tab_per_department(client):
    await _seed()
    client.cookies.set("survey_admin_session", "1")

    r = await client.get("/admin/survey/export-full-report")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert ".xlsx" in r.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames[0] == "Overview"
    assert wb.sheetnames[1] == "Charts"
    assert LAW in wb.sheetnames
    assert COMMERCE in wb.sheetnames

    overview = wb["Overview"]
    header = [c.value for c in overview[4]]
    assert header == ["Department", "Registered", "Baseline done", "Baseline pending",
                      "Post survey done", "Post survey pending",
                      "Deeksharambh done", "Deeksharambh pending"]

    law_ws = wb[LAW]
    student_header = [c.value for c in law_ws[4]]
    assert student_header == [
        "Name", "Email", "Level", "Registered", "Baseline done", "Baseline date",
        "Post survey done", "Post survey date", "Deeksharambh done", "Deeksharambh date"]
    body = [[c.value for c in row] for row in law_ws.iter_rows(min_row=5)]
    assert len(body) == 3
    emails = {row[1] for row in body}
    assert emails == {"law1@x.com", "law2@x.com", "law3@x.com"}
