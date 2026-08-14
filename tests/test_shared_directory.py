"""
The overall department directory — one shareable page covering every
department, with each department's own analysis and export links inside it.
"""

from __future__ import annotations

from datetime import datetime, timezone
from typing import AsyncIterator

import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from mongomock_motor import AsyncMongoMockClient

from app import db

LAW = "Department of Law"
COM = "Department of Commerce"


@pytest_asyncio.fixture
async def app_with_mock():
    db._set_client_for_tests(AsyncMongoMockClient())
    try:
        from app.main import app
        await db.init_indexes(allow_duplicate_email=True)
        yield app
    finally:
        db._reset_clients_for_tests()


@pytest_asyncio.fixture
async def client(app_with_mock) -> AsyncIterator[AsyncClient]:
    transport = ASGITransport(app=app_with_mock)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        yield ac


async def _seed() -> None:
    from app.hacri_e2_compat import SCHEMA

    now = datetime.now(timezone.utc)
    people = [
        ("a@example.com", LAW, db.STATUS_POST_DONE, True, True),   # mailed + clicked
        ("b@example.com", LAW, db.STATUS_PRE_DONE, True, False),   # mailed, no click
        ("c@example.com", LAW, None, False, False),
        ("d@example.com", COM, db.STATUS_PRE_DONE, False, False),
    ]
    for email, program, status, mailed, clicked in people:
        doc = {
            "email": email, "name": email[0].upper(), "program": program,
            "ug_or_pg": "ug", "status": status, "created_at": now,
            "pre_submitted_at": now if status else None,
            "post_submitted_at": now if status == db.STATUS_POST_DONE else None,
        }
        if mailed:
            doc["post_reminder_sent_at"] = now
        if clicked:
            # Clicked before submitting, so it counts as "filled after mail".
            doc["reminder_clicked_at"] = now.replace(year=now.year - 1)
        await db.get_db()["users"].insert_one(doc)

        if status:
            await db.get_db()["pre_responses"].insert_one(
                {"email": email, "submitted_at": now, "fields": {k: 3 for k in SCHEMA}})
        if status == db.STATUS_POST_DONE:
            await db.get_db()["post_responses"].insert_one(
                {"email": email, "submitted_at": now, "fields": {k: 4 for k in SCHEMA}})


def _token() -> str:
    from app.routes.shared_analysis import get_directory_token
    return get_directory_token()


@pytest.mark.asyncio
async def test_directory_needs_the_token(client: AsyncClient):
    await _seed()

    assert (await client.get("/shared/departments")).status_code == 422   # token missing
    assert (await client.get("/shared/departments?token=nope")).status_code == 403
    assert (await client.get(f"/shared/departments?token={_token()}")).status_code == 200


@pytest.mark.asyncio
async def test_directory_shows_counts_for_every_department(client: AsyncClient):
    await _seed()

    page = (await client.get(f"/shared/departments?token={_token()}")).text

    assert LAW in page and COM in page
    assert "Overall (All Departments)" in page
    # Column headings the admin asked for.
    for heading in ("Registered", "Baseline", "Post", "pending", "Reminders", "Clicked"):
        assert heading in page


@pytest.mark.asyncio
async def test_directory_counts_are_right(client: AsyncClient):
    """Check the numbers through the route's own context, not the markup."""
    from app.db import get_dept_analysis_data, get_email_notification_stats

    await _seed()
    analysis = await get_dept_analysis_data()
    mail = {row["dept"]: row for row in await get_email_notification_stats()}

    law = next(d for d in analysis["departments"] if d["dept"] == LAW)
    assert law["registered"] == 3
    assert law["pre_done"] == 2          # one post_done counts as pre done too
    assert law["post_done"] == 1
    assert mail[LAW]["post_sent"] == 2
    assert mail[LAW]["clicked"] == 1
    assert mail[LAW]["completed_after"] == 1

    assert analysis["overall"]["registered"] == 4


@pytest.mark.asyncio
async def test_every_row_carries_analysis_and_excel_links(client: AsyncClient):
    await _seed()
    page = (await client.get(f"/shared/departments?token={_token()}")).text

    from app.routes.shared_analysis import get_dept_token
    for dept in (LAW, COM, "Overall"):
        for kind in ("pre", "post"):
            token = get_dept_token(dept, kind)
            assert f"token={token}&type={kind}" in page, f"{dept}/{kind} analysis link"
        assert "export-excel" in page
        assert "download-ppt" in page


@pytest.mark.asyncio
async def test_department_links_from_the_directory_actually_open(client: AsyncClient):
    await _seed()
    from app.routes.shared_analysis import get_dept_token

    token = get_dept_token(LAW, "pre")
    page = await client.get(f"/shared/analysis?dept={LAW}&token={token}&type=pre")
    assert page.status_code == 200

    excel = await client.get(f"/shared/analysis/export-excel?dept={LAW}&token={token}&type=pre")
    assert excel.status_code == 200
    assert excel.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml")
    assert len(excel.content) > 0

    # A token for one department must not open another.
    other = await client.get(f"/shared/analysis?dept={COM}&token={token}&type=pre")
    assert other.status_code == 403


@pytest.mark.asyncio
async def test_overall_excel_covers_every_department(client: AsyncClient):
    """'Overall' is not a department name — it must not export an empty file."""
    await _seed()
    from app.routes.shared_analysis import get_dept_token

    token = get_dept_token("Overall", "pre")
    resp = await client.get(f"/shared/analysis/export-excel?dept=Overall&token={token}&type=pre")
    assert resp.status_code == 200

    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.content))

    # Sheet 1: every department, side by side.
    assert wb.sheetnames[0] == "Department Breakdown"
    breakdown = list(wb["Department Breakdown"].iter_rows(values_only=True))
    flat = [[c for c in row] for row in breakdown]
    depts = {row[0] for row in flat if row and isinstance(row[0], str)}
    assert LAW in depts and COM in depts
    assert "ALL DEPARTMENTS" in depts          # totals line

    law_row = next(r for r in flat if r and r[0] == LAW)
    assert law_row[1:4] == [3, 2, 1]           # registered / baseline / post
    totals_row = next(r for r in flat if r and r[0] == "ALL DEPARTMENTS")
    assert totals_row[1] == 4                  # every registered student

    # Then the rosters: everyone who has filled the baseline, and everyone
    # registered who still owes it.
    filled = list(wb["Filled - Pre Survey"].iter_rows(values_only=True))
    emails = {c for row in filled for c in row if isinstance(c, str) and "@example.com" in c}
    assert emails == {"a@example.com", "b@example.com", "d@example.com"}

    pending = list(wb["Registered - Pre Not Filled"].iter_rows(values_only=True))
    pending_emails = {c for row in pending for c in row
                      if isinstance(c, str) and "@example.com" in c}
    assert pending_emails == {"c@example.com"}


async def test_single_department_excel_has_no_breakdown_sheet(client: AsyncClient):
    """The breakdown only makes sense on the all-departments export."""
    await _seed()
    from app.routes.shared_analysis import get_dept_token

    token = get_dept_token(LAW, "pre")
    resp = await client.get(f"/shared/analysis/export-excel?dept={LAW}&token={token}&type=pre")
    assert resp.status_code == 200

    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["Summary", "Registered - Pre Not Filled",
                             "Filled - Pre Survey"]


@pytest.mark.asyncio
async def test_admin_links_api_serves_the_directory_url(client: AsyncClient):
    await _seed()
    client.cookies.set("survey_admin_session", "1")

    data = (await client.get("/admin/api/survey/post-links")).json()
    assert "/shared/departments?token=" in data["directory_url"]

    # The link the admin hands out must be the one that opens.
    path = data["directory_url"].split("http://test", 1)[-1]
    assert (await client.get(path)).status_code == 200


@pytest.mark.asyncio
async def test_each_row_can_draft_a_mail(client: AsyncClient):
    """Pre drafts carry the analysis + Excel; post drafts also carry the
    student survey link."""
    await _seed()
    page = (await client.get(f"/shared/departments?token={_token()}")).text

    # Two mail buttons per row — one for the baseline, one for the post survey.
    assert page.count('data-mail="pre"') == 3      # Overall + two departments
    assert page.count('data-mail="post"') == 3

    # The post draft needs the student-facing link; the baseline one does not.
    assert 'data-student-path="/post/department-of-law"' in page
    assert 'data-student-path="/post/all"' in page

    from app.routes.shared_analysis import get_dept_token
    assert f'data-token="{get_dept_token(LAW, "pre")}"' in page
    assert f'data-token="{get_dept_token(LAW, "post")}"' in page

    # Department names travel as data attributes, so a quote or bracket in a
    # name can never break the markup the way an inline onclick would.
    assert 'onclick="draftMail' not in page


@pytest.mark.asyncio
async def test_compose_dialog_is_available_on_both_shared_pages(client: AsyncClient):
    """The mail dialog asks for a subject and message before anything opens."""
    await _seed()
    from app.routes.shared_analysis import get_dept_token

    directory = (await client.get(f"/shared/departments?token={_token()}")).text
    analysis = (await client.get(
        f"/shared/analysis?dept={LAW}&token={get_dept_token(LAW, 'post')}&type=post")).text

    for page, where in ((directory, "directory"), (analysis, "analysis page")):
        assert 'id="mc-overlay"' in page, f"compose dialog missing from the {where}"
        assert 'id="mc-subject"' in page, f"subject field missing from the {where}"
        assert 'id="mc-message"' in page, f"message field missing from the {where}"
        # Both flavours on offer.
        assert 'data-mode="plain"' in page and 'data-mode="html"' in page

    # The analysis page carries its own button and its department's links.
    assert 'id="btn-mail-report"' in analysis
    assert '"/post/department-of-law"' in analysis


@pytest.mark.asyncio
async def test_overall_analysis_page_links_the_all_departments_survey(client: AsyncClient):
    await _seed()
    from app.routes.shared_analysis import get_dept_token

    page = (await client.get(
        f"/shared/analysis?dept=Overall&token={get_dept_token('Overall', 'post')}&type=post")).text
    assert '"/post/all"' in page


@pytest.mark.asyncio
async def test_mail_exports_are_recorded_against_the_department(client: AsyncClient):
    """Drafting a mail records who it was addressed to, and how many times."""
    await _seed()
    from app.routes.shared_analysis import get_dept_token

    token = get_dept_token(LAW, "post")
    payload = {"dept": LAW, "token": token, "type": "post", "format": "html",
               "via": "gmail", "subject": "Report",
               "to": ["HOD@jainuniversity.ac.in", "dean@jainuniversity.ac.in"]}

    first = await client.post("/shared/analysis/mail-log", json=payload)
    assert first.status_code == 200
    history = first.json()["history"]
    assert history["total"] == 1
    # Addresses are normalised, so the same person is not counted twice.
    assert {r["email"] for r in history["recipients"]} == {
        "hod@jainuniversity.ac.in", "dean@jainuniversity.ac.in"}

    # Mailing the same person again increments their count.
    payload["to"] = ["hod@jainuniversity.ac.in"]
    second = await client.post("/shared/analysis/mail-log", json=payload)
    counts = {r["email"]: r["count"] for r in second.json()["history"]["recipients"]}
    assert counts["hod@jainuniversity.ac.in"] == 2
    assert counts["dean@jainuniversity.ac.in"] == 1
    assert second.json()["history"]["total"] == 2

    # …and it shows on that department's analysis page.
    page = (await client.get(
        f"/shared/analysis?dept={LAW}&token={token}&type=post")).text
    assert "Mail history" in page
    assert "hod@jainuniversity.ac.in" in page


@pytest.mark.asyncio
async def test_mail_log_needs_a_valid_department_token(client: AsyncClient):
    await _seed()
    from app.routes.shared_analysis import get_dept_token

    bad = await client.post("/shared/analysis/mail-log", json={
        "dept": LAW, "token": "nonsense", "type": "post", "to": ["x@y.com"]})
    assert bad.status_code == 403

    # A token for one department cannot log against another.
    crossed = await client.post("/shared/analysis/mail-log", json={
        "dept": COM, "token": get_dept_token(LAW, "post"), "type": "post",
        "to": ["x@y.com"]})
    assert crossed.status_code == 403

    # Nothing was written.
    page = (await client.get(
        f"/shared/analysis?dept={COM}&token={get_dept_token(COM, 'post')}&type=post")).text
    assert "has not been mailed yet" in page


@pytest.mark.asyncio
async def test_a_draft_with_no_recipients_still_counts_as_an_export(client: AsyncClient):
    await _seed()
    from app.routes.shared_analysis import get_dept_token

    resp = await client.post("/shared/analysis/mail-log", json={
        "dept": LAW, "token": get_dept_token(LAW, "pre"), "type": "pre",
        "format": "plain", "via": "copy-plain", "subject": "S", "to": []})
    history = resp.json()["history"]
    assert history["total"] == 1
    assert history["recipients"] == []


@pytest.mark.asyncio
async def test_mail_drafts_carry_the_department_figures(client: AsyncClient):
    """The mail should show the numbers, not only link to them."""
    await _seed()
    from app.routes.shared_analysis import get_dept_token

    page = (await client.get(f"/shared/departments?token={_token()}")).text
    # Every Mail button hands its row's figures to the compose dialog.
    assert 'data-registered=' in page and 'data-post-pending=' in page
    assert 'data-lit-pre=' in page and 'data-read-post=' in page

    analysis = (await client.get(
        f"/shared/analysis?dept={LAW}&token={get_dept_token(LAW, 'post')}&type=post")).text
    # Law: 3 registered, 2 baseline done, 1 post done, 1 still pending.
    assert '"registered": 3' in analysis or "registered: 3" in analysis
    assert "pre_done: 2" in analysis
    assert "post_done: 1" in analysis


@pytest.mark.asyncio
async def test_excel_counts_everyone_registered_not_only_the_finishers(client: AsyncClient):
    """The workbook's registered figure has to match the report page.

    It used to count only the students already in the sheet, so a department
    with 59 registered and 50 finished exported "50 registered".
    """
    await _seed()
    from app.routes.shared_analysis import get_dept_token

    token = get_dept_token(LAW, "pre")
    resp = await client.get(f"/shared/analysis/export-excel?dept={LAW}&token={token}&type=pre")

    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.content))

    summary = {row[0]: row[1] for row in wb["Summary"].iter_rows(values_only=True)
               if row and isinstance(row[0], str)}
    assert summary["Total Registered Students"] == 3     # a, b and c
    assert summary["Baseline (Pre) Survey Completed"] == 2
    assert summary["Baseline (Pre) Survey Pending"] == 1


@pytest.mark.asyncio
async def test_the_roster_sheets_carry_only_the_five_columns(client: AsyncClient):
    await _seed()
    from app.routes.shared_analysis import get_dept_token

    token = get_dept_token(LAW, "pre")
    resp = await client.get(f"/shared/analysis/export-excel?dept={LAW}&token={token}&type=pre")

    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.content))

    expected = ["Name", "Email", "Department", "Level", "Education Type"]
    for sheet in ("Registered - Pre Not Filled", "Filled - Pre Survey"):
        ws = wb[sheet]
        assert [ws.cell(4, c).value for c in range(1, ws.max_column + 1)] == expected
        # No score columns anywhere in the sheet.
        text = " ".join(str(c.value) for row in ws.iter_rows() for c in row)
        assert "Literacy" not in text and "Quadrant" not in text

    assert {ws.cell(r, 2).value for r in range(5, wb["Filled - Pre Survey"].max_row + 1)} == \
        {"a@example.com", "b@example.com"}
    assert wb["Registered - Pre Not Filled"].cell(5, 2).value == "c@example.com"


@pytest.mark.asyncio
async def test_the_post_export_splits_on_the_post_survey(client: AsyncClient):
    await _seed()
    from app.routes.shared_analysis import get_dept_token

    token = get_dept_token(LAW, "post")
    resp = await client.get(f"/shared/analysis/export-excel?dept={LAW}&token={token}&type=post")

    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.content))

    assert wb.sheetnames == ["Summary", "Registered - Post Not Filled",
                             "Filled - Post Survey"]
    assert wb["Filled - Post Survey"].cell(5, 2).value == "a@example.com"
    pending = {wb["Registered - Post Not Filled"].cell(r, 2).value
               for r in range(5, wb["Registered - Post Not Filled"].max_row + 1)}
    assert pending == {"b@example.com", "c@example.com"}
