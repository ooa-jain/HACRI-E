"""The Excel download the impact page now offers, through the real route."""
from __future__ import annotations
import io
from datetime import datetime, timedelta, timezone
from typing import AsyncIterator

import pytest, pytest_asyncio
from httpx import ASGITransport, AsyncClient
from mongomock_motor import AsyncMongoMockClient
from openpyxl import load_workbook

from app import db
from app.routes.shared_analysis import get_orientation_token, get_vibe_token

BIG = "Department of Commerce"
TINY = "Department of Art and Design"


@pytest_asyncio.fixture
async def client() -> AsyncIterator[AsyncClient]:
    db._set_client_for_tests(AsyncMongoMockClient())
    try:
        from app.main import app
        await db.init_indexes(allow_duplicate_email=True)
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as ac:
            yield ac
    finally:
        db._reset_clients_for_tests()


async def _seed(big: int = 14, tiny: int = 2) -> None:
    now = datetime.now(timezone.utc)
    for dept, count in ((BIG, big), (TINY, tiny)):
        for i in range(count):
            email = f"{dept[-4:]}{i}@x.com".replace(" ", "")
            await db.get_db()["users"].insert_one({
                "email": email, "name": f"S{i}", "program": dept, "ug_or_pg": "ug",
                "location": "Bangalore", "status": db.STATUS_PRE_DONE,
                "created_at": now, "pre_submitted_at": now,
                "orientation_submitted": True})
            await db.get_db()["orientation_responses"].insert_one({
                "email": email, "name": f"S{i}", "submitted_at": now - timedelta(hours=i),
                "data": {"location": "📍 Bangalore", "q2": 8, "q29": 7, "q34": 9,
                         "q16": 4, "q37": ["🌉 Bridge course sessions"],
                         "q38": ["😴 Too many long sitting sessions"],
                         "q39": ["🎯 More interactive workshops"],
                         "q11": ["🚶 Campus Tour"]}})


def _link(campus="Bangalore"):
    return {"campus": campus, "token": get_orientation_token(campus)}


@pytest.mark.asyncio
async def test_the_workbook_downloads_and_carries_the_report(client):
    await _seed()
    r = await client.get("/shared/orientation/excel", params=_link())
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert ".xlsx" in r.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["Overview", "Departments", "Feedback", "All questions"]

    # Both departments are listed — coverage is the point of the sheet.
    depts = [row[0] for row in wb["Departments"].iter_rows(min_row=6, values_only=True)]
    assert BIG in depts and TINY in depts

    # The one with two respondents keeps its counts and loses its scores.
    for row in wb["Departments"].iter_rows(min_row=6, values_only=True):
        if row[0] == TINY:
            assert row[1] == 2                       # answered
            assert "withheld" in str(row[5])         # vibe
        if row[0] == BIG:
            assert row[1] == 14
            assert row[5] == 8.0

    # Every feedback question reaches the workbook with its own denominator.
    questions = {row[0] for row in wb["Feedback"].iter_rows(min_row=6, values_only=True)}
    assert {"Keep next year", "Stop next year", "Introduce next year"} <= questions


@pytest.mark.asyncio
async def test_a_department_link_downloads_only_its_own_workbook(client):
    await _seed()
    params = {"campus": "Bangalore", "dept": BIG,
              "token": get_orientation_token("Bangalore", BIG)}
    wb = load_workbook(io.BytesIO((await client.get(
        "/shared/orientation/excel", params=params)).content))
    depts = [row[0] for row in wb["Departments"].iter_rows(min_row=6, values_only=True)]
    assert depts == [BIG]

    # And a token minted for one department opens nobody else's.
    bad = {**params, "dept": TINY}
    assert (await client.get("/shared/orientation/excel", params=bad)).status_code == 403


@pytest.mark.asyncio
async def test_the_impact_page_offers_the_deck_and_the_workbook(client):
    await _seed()
    page = (await client.get("/shared/impact",
                             params={"campus": "", "token": get_vibe_token("")})).text
    assert "/shared/orientation/excel?" in page
    assert "/shared/orientation/ppt?" in page
    assert ">Excel</a>" in page and ">Deck</a>" in page
