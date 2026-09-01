"""
school_export.py — the school workbook.

One file that opens the way the pages read: the whole university first, then
every department on one sheet, then a tab per school carrying the departments
inside it. Baseline and post sit side by side in every table with the change
between them, because the change is the point of running the survey twice.
"""
from __future__ import annotations

import io
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY = "1B2A4A"
GOLD = "C9A84C"
MIST = "F5F6FA"
GREEN = "15803D"
ROSE = "9F1239"


def _styles() -> dict:
    return {
        "navy": PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid"),
        "gold": PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid"),
        "gray": PatternFill(start_color=MIST, end_color=MIST, fill_type="solid"),
        "white_font": Font(name="Segoe UI", size=11, bold=True, color="FFFFFF"),
        "title_font": Font(name="Segoe UI", size=16, bold=True, color=NAVY),
        "sub_font": Font(name="Segoe UI", size=10, color="64748B"),
        "bold_font": Font(name="Segoe UI", size=11, bold=True),
        "normal_font": Font(name="Segoe UI", size=11),
        "up_font": Font(name="Segoe UI", size=11, bold=True, color=GREEN),
        "down_font": Font(name="Segoe UI", size=11, bold=True, color=ROSE),
        "border": Border(
            left=Side(style="thin", color="DDDDDD"),
            right=Side(style="thin", color="DDDDDD"),
            top=Side(style="thin", color="DDDDDD"),
            bottom=Side(style="thin", color="DDDDDD"),
        ),
    }


def _safe_sheet_name(name: str, used: set[str]) -> str:
    """A legal, unique Excel tab name. Excel forbids \\ / * ? : [ ] and caps
    names at 31 characters, which several school names exceed."""
    cleaned = re.sub(r"[\\/*?:\[\]]", "", name or "").strip() or "Sheet"
    candidate = cleaned[:31]
    n = 2
    while candidate.lower() in used:
        suffix = f" ({n})"
        candidate = cleaned[: 31 - len(suffix)] + suffix
        n += 1
    used.add(candidate.lower())
    return candidate


def _delta(post, pre):
    """Post minus baseline, or None when either side has no answers yet."""
    if post is None or pre is None:
        return None
    return round(float(post) - float(pre), 2)


def _value(v):
    """Excel shows a blank rather than the string "None"."""
    return "" if v is None else v


def _header(ws, st, title: str, subtitle: str, headers: list[str],
            widths: list[int]) -> int:
    """Title, subtitle, then the header band. Returns the first data row."""
    ws["A1"] = title
    ws["A1"].font = st["title_font"]
    ws["A2"] = subtitle
    ws["A2"].font = st["sub_font"]

    row = 4
    for i, name in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=name)
        cell.fill = st["navy"]
        cell.font = st["white_font"]
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = st["border"]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    return row + 1


def _write_row(ws, st, row_idx: int, values: list, *, bold_first=False,
               delta_cols: tuple[int, ...] = ()) -> None:
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col, value=_value(value))
        cell.border = st["border"]
        cell.font = st["normal_font"]
        if bold_first and col == 1:
            cell.font = st["bold_font"]
        if col > 1:
            cell.alignment = Alignment(horizontal="center")
        # A change is worth seeing as up or down without reading the number.
        if col in delta_cols and isinstance(value, (int, float)):
            cell.font = st["up_font"] if value > 0 else (
                st["down_font"] if value < 0 else st["normal_font"])
            cell.number_format = "+0.00;-0.00;0.00"


SCHOOL_HEADERS = [
    "School", "Departments", "Registered", "Baseline done", "Post done",
    "Baseline pending", "Post pending", "Total submissions", "Filled recently",
    "Last submission", "Avg literacy (baseline)", "Avg literacy (post)",
    "Literacy change", "Avg readiness (baseline)", "Avg readiness (post)",
    "Readiness change",
]
SCHOOL_WIDTHS = [42, 12, 12, 14, 12, 15, 13, 16, 14, 20, 18, 16, 15, 19, 17, 16]

DEPT_HEADERS = [
    "School", "Department", "Registered", "Baseline done", "Post done",
    "Baseline pending", "Post pending", "Total submissions",
    "Avg literacy (baseline)", "Avg literacy (post)", "Literacy change",
    "Avg readiness (baseline)", "Avg readiness (post)", "Readiness change",
]
DEPT_WIDTHS = [38, 42, 12, 14, 12, 15, 13, 16, 18, 16, 15, 19, 17, 16]


def _school_values(s: dict) -> list:
    return [
        s["school"], s["dept_count"], s["registered"], s["pre_done"],
        s["post_done"], s["pre_pending"], s["post_pending"],
        s["pre_done"] + s["post_done"], s["recent_total"],
        s.get("last_submission") or "—",
        s.get("avg_lit_pre"), s.get("avg_lit_post"),
        _delta(s.get("avg_lit_post"), s.get("avg_lit_pre")),
        s.get("avg_read_pre"), s.get("avg_read_post"),
        _delta(s.get("avg_read_post"), s.get("avg_read_pre")),
    ]


def _dept_values(school: str, d: dict) -> list:
    return [
        school, d["dept"], d["registered"], d["pre_done"], d["post_done"],
        d["pre_pending"], d["post_pending"], d["pre_done"] + d["post_done"],
        d.get("avg_lit_pre"), d.get("avg_lit_post"),
        _delta(d.get("avg_lit_post"), d.get("avg_lit_pre")),
        d.get("avg_read_pre"), d.get("avg_read_post"),
        _delta(d.get("avg_read_post"), d.get("avg_read_pre")),
    ]


def _totals_row(ws, st, row_idx: int, label: str, values: list) -> None:
    for col, value in enumerate([label] + values, start=1):
        cell = ws.cell(row=row_idx, column=col, value=_value(value))
        cell.fill = st["gray"]
        cell.font = st["bold_font"]
        cell.border = st["border"]
        if col > 1:
            cell.alignment = Alignment(horizontal="center")


def _all_schools_sheet(wb, st, data: dict, generated_at: str) -> None:
    ov = data["overall"]
    ws = wb.active
    ws.title = "All Schools"

    row = _header(
        ws, st, "HACRI-E — All Schools",
        f"Baseline and post side by side, with the change between them. "
        f"{ov['schools_listed']} schools · {ov['school_count']} with students · "
        f"generated {generated_at}",
        SCHOOL_HEADERS, SCHOOL_WIDTHS)

    for school in data["schools"]:
        _write_row(ws, st, row, _school_values(school), bold_first=True,
                   delta_cols=(13, 16))
        row += 1

    _totals_row(ws, st, row, "TOTAL — all schools", [
        sum(s["dept_count"] for s in data["schools"]),
        ov["registered"], ov["pre_done"], ov["post_done"],
        max(0, ov["registered"] - ov["pre_done"]),
        max(0, ov["pre_done"] - ov["post_done"]),
        ov["submissions"], ov["recent_total"], "",
        ov.get("avg_lit_pre"), ov.get("avg_lit_post"),
        _delta(ov.get("avg_lit_post"), ov.get("avg_lit_pre")),
        ov.get("avg_read_pre"), ov.get("avg_read_post"),
        _delta(ov.get("avg_read_post"), ov.get("avg_read_pre")),
    ])


def _all_departments_sheet(wb, st, data: dict, generated_at: str) -> None:
    """Every department in the university, on one sheet, named by its school."""
    ws = wb.create_sheet("Departments")
    total = sum(len(s["departments"]) for s in data["schools"])

    row = _header(
        ws, st, "HACRI-E — Every Department",
        f"{total} departments across {data['overall']['schools_listed']} schools, "
        f"sorted by school. Generated {generated_at}",
        DEPT_HEADERS, DEPT_WIDTHS)

    for school in data["schools"]:
        for dept in school["departments"]:
            _write_row(ws, st, row, _dept_values(school["school"], dept),
                       delta_cols=(11, 14))
            row += 1


def _one_school_sheet(wb, st, school: dict, used: set[str],
                      generated_at: str) -> None:
    ws = wb.create_sheet(_safe_sheet_name(school["school"], used))

    row = _header(
        ws, st, school["school"],
        f"{school['dept_count']} department"
        f"{'' if school['dept_count'] == 1 else 's'} · "
        f"{school['registered']} registered · {school['pre_done']} baseline · "
        f"{school['post_done']} post · generated {generated_at}",
        DEPT_HEADERS[1:], DEPT_WIDTHS[1:])

    for dept in school["departments"]:
        _write_row(ws, st, row, _dept_values("", dept)[1:], bold_first=True,
                   delta_cols=(10, 13))
        row += 1

    if not school["departments"]:
        ws.cell(row=row, column=1,
                value="No department registers under this school yet.")
        ws.cell(row=row, column=1).font = st["sub_font"]
        row += 1

    _totals_row(ws, st, row, f"TOTAL — {school['school']}", [
        school["registered"], school["pre_done"], school["post_done"],
        school["pre_pending"], school["post_pending"],
        school["pre_done"] + school["post_done"],
        school.get("avg_lit_pre"), school.get("avg_lit_post"),
        _delta(school.get("avg_lit_post"), school.get("avg_lit_pre")),
        school.get("avg_read_pre"), school.get("avg_read_post"),
        _delta(school.get("avg_read_post"), school.get("avg_read_pre")),
    ])


def generate_schools_excel(data: dict, *, generated_at: str = "") -> bytes:
    """The whole university: all schools, every department, then a tab each.

    Sheet order is the order the question gets asked — the university, then
    the departments, then each school on its own — so the file can be sent to
    a dean who only opens their own tab.
    """
    wb = Workbook()
    st = _styles()

    _all_schools_sheet(wb, st, data, generated_at)
    _all_departments_sheet(wb, st, data, generated_at)

    used: set[str] = {"all schools", "departments"}
    for school in data["schools"]:
        # A school nobody registered under has nothing to put on a tab; it is
        # already listed on the first sheet with its zeros.
        if school["registered"]:
            _one_school_sheet(wb, st, school, used, generated_at)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_one_school_excel(school: dict, *, generated_at: str = "") -> bytes:
    """One school on its own: its departments, and its own total."""
    wb = Workbook()
    st = _styles()
    wb.active.title = _safe_sheet_name(school["school"], set())
    ws = wb.active

    row = _header(
        ws, st, school["school"],
        f"{school['dept_count']} department"
        f"{'' if school['dept_count'] == 1 else 's'} · "
        f"{school['registered']} registered · {school['pre_done']} baseline · "
        f"{school['post_done']} post · generated {generated_at}",
        DEPT_HEADERS[1:], DEPT_WIDTHS[1:])

    for dept in school["departments"]:
        _write_row(ws, st, row, _dept_values("", dept)[1:], bold_first=True,
                   delta_cols=(10, 13))
        row += 1

    if not school["departments"]:
        ws.cell(row=row, column=1,
                value="No department registers under this school yet.")
        ws.cell(row=row, column=1).font = st["sub_font"]
        row += 1

    _totals_row(ws, st, row, f"TOTAL — {school['school']}", [
        school["registered"], school["pre_done"], school["post_done"],
        school["pre_pending"], school["post_pending"],
        school["pre_done"] + school["post_done"],
        school.get("avg_lit_pre"), school.get("avg_lit_post"),
        _delta(school.get("avg_lit_post"), school.get("avg_lit_pre")),
        school.get("avg_read_pre"), school.get("avg_read_post"),
        _delta(school.get("avg_read_post"), school.get("avg_read_pre")),
    ])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
