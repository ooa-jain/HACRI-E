"""
Excel export helper using openpyxl.

The shared department workbook is a roster, not a scoresheet: who is
registered and still owes the survey, and who has filled it. Scores live in
the report itself and in the admin's own custom export, so this file stays
readable by the people who chase students.
"""
from __future__ import annotations
import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

# The only columns the roster sheets carry.
ROSTER_HEADERS = ["Name", "Email", "Department", "Level", "Education Type"]
_COLUMN_WIDTHS = [26, 32, 38, 10, 22]

NAVY = "1B2A4A"
GOLD = "C9A84C"
MIST = "F5F6FA"


def _styles():
    return {
        "navy": PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid"),
        "gold": PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid"),
        "gray": PatternFill(start_color=MIST, end_color=MIST, fill_type="solid"),
        "white_font": Font(name="Segoe UI", size=11, bold=True, color="FFFFFF"),
        "title_font": Font(name="Segoe UI", size=16, bold=True, color="1B2A4A"),
        "bold_font": Font(name="Segoe UI", size=11, bold=True),
        "normal_font": Font(name="Segoe UI", size=11),
        "border": Border(
            left=Side(style="thin", color="DDDDDD"),
            right=Side(style="thin", color="DDDDDD"),
            top=Side(style="thin", color="DDDDDD"),
            bottom=Side(style="thin", color="DDDDDD"),
        ),
    }


def _has_pre(user: dict) -> bool:
    return user.get("status") in ("pre_done", "post_done")


def _has_post(user: dict) -> bool:
    return user.get("status") == "post_done"


def _roster_sheet(wb, title: str, subtitle: str, users: list[dict], st, fill) -> None:
    """One tab: the five columns, nothing else."""
    ws = wb.create_sheet(title[:31])
    ws.views.sheetView[0].showGridLines = True

    ws["A1"] = title
    ws["A1"].font = st["title_font"]
    ws["A2"] = subtitle
    ws["A2"].font = st["bold_font"]

    header_row = 4
    for col_idx, text in enumerate(ROSTER_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=text)
        cell.font = st["white_font"]
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = st["border"]
    ws.row_dimensions[header_row].height = 24

    for row_idx, user in enumerate(users, start=header_row + 1):
        values = [
            user.get("name", ""),
            user.get("email", ""),
            user.get("program", "") or "—",
            (user.get("ug_or_pg", "ug") or "ug").upper(),
            user.get("education_type", "") or "—",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = st["normal_font"]
            cell.border = st["border"]
            cell.alignment = Alignment(horizontal="center" if col_idx == 4 else "left")

    if not users:
        cell = ws.cell(row=header_row + 1, column=1, value="Nobody in this list")
        cell.font = st["normal_font"]

    for col_idx, width in enumerate(_COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def _fill_department_summary_sheet(ws, st, summary: dict, generated_at: str = "") -> None:
    """Write the Registered/Pre/Post/Deeksharambh table into `ws`.

    Shared by the standalone single-tab export and the "Overview" tab of the
    full report, so the two can never drift apart on numbers or layout.
    """
    ws["A1"] = "Deeksharambh 2026 — Department Summary"
    ws["A1"].font = st["title_font"]
    ws["A2"] = f"Registered, baseline, post survey and Deeksharambh, per department."                + (f"  Generated {generated_at}." if generated_at else "")
    ws["A2"].font = st["normal_font"]

    headers = ["Department", "Registered",
               "Baseline done", "Baseline pending",
               "Post survey done", "Post survey pending",
               "Deeksharambh done", "Deeksharambh pending"]
    header_row = 4
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=text)
        cell.font = st["white_font"]
        cell.fill = st["navy"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = st["border"]
    ws.row_dimensions[header_row].height = 32

    rows = summary.get("departments") or []
    totals = summary.get("totals") or {}

    def _row(row_idx, dept, r, *, bold=False, fill=None):
        values = [
            dept, r.get("registered", 0),
            r.get("pre_done", 0), r.get("pre_pending", 0),
            r.get("post_done", 0), r.get("post_pending", 0),
            r.get("orientation_done", 0), r.get("orientation_pending", 0),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = st["bold_font"] if bold else st["normal_font"]
            cell.border = st["border"]
            cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "center")
            if fill:
                cell.fill = fill

    # The cohort total leads, so a reader gets the whole picture before the
    # per-department breakdown — and it stays visible under a frozen header
    # even after the sheet is sorted, since freeze_panes below sits under it.
    _row(header_row + 1, "All departments", totals, bold=True, fill=st["gray"])
    for i, row in enumerate(rows, start=header_row + 2):
        _row(i, row["dept"], row)

    if not rows:
        cell = ws.cell(row=header_row + 2, column=1, value="No departments registered yet")
        cell.font = st["normal_font"]

    widths = [40, 12, 14, 16, 17, 19, 18, 20]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = ws.cell(row=header_row + 2, column=1)
    ws.auto_filter.ref = f"A{header_row}:H{header_row + len(rows) + 1}"


def generate_department_summary_excel(summary: dict, *, generated_at: str = "") -> bytes:
    """One sheet, one row per department: Registered, Pre, Post, Deeksharambh.

    The single-tab export the overview page offers — every department's four
    headline counts side by side, so a reader can scan or sort the whole
    cohort without opening a report per department. `summary` is a
    `department_registration_summary()` result; the numbers are not
    recomputed here.
    """
    st = _styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "Department Summary"
    _fill_department_summary_sheet(ws, st, summary, generated_at)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_cohort_excel(
    dept_name: str,
    users_list: list[dict],
    survey_type: str = "pre",
    dept_rows: list[dict] | None = None,
) -> bytes:
    """Roster workbook for one department — or for all of them.

    `users_list` is every registered student in scope, filled or not: the
    counts here are meant to match the report page, and they cannot if the
    caller has already dropped the students who never answered.

    Three sheets — a summary, who still owes the survey, and who has filled
    it. `dept_rows` adds the per-department breakdown used by the
    all-departments export.
    """
    label = "Post" if survey_type == "post" else "Pre"
    done = _has_post if survey_type == "post" else _has_pre

    filled = [u for u in users_list if done(u)]
    not_filled = [u for u in users_list if not done(u)]

    st = _styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.views.sheetView[0].showGridLines = True

    ws["A1"] = "HACRI-E2 Department Cohort"
    ws["A1"].font = st["title_font"]
    ws["A2"] = f"Department: {dept_name or 'All Departments'}"
    ws["A2"].font = st["bold_font"]

    stats = [
        ("Total Registered Students", len(users_list)),
        ("Baseline (Pre) Survey Completed", sum(1 for u in users_list if _has_pre(u))),
        ("Baseline (Pre) Survey Pending", sum(1 for u in users_list if not _has_pre(u))),
        ("Post-Workshop Survey Completed", sum(1 for u in users_list if _has_post(u))),
        ("Post-Workshop Survey Pending", sum(1 for u in users_list if not _has_post(u))),
        (f"In this workbook — {label} filled", len(filled)),
        (f"In this workbook — {label} not filled", len(not_filled)),
    ]

    for col_idx, text in enumerate(["Metric", "Count"], start=1):
        cell = ws.cell(row=4, column=col_idx, value=text)
        cell.font = st["white_font"]
        cell.fill = st["navy"]
        cell.alignment = Alignment(horizontal="center")
        cell.border = st["border"]

    for row_idx, (metric, value) in enumerate(stats, start=5):
        name_cell = ws.cell(row=row_idx, column=1, value=metric)
        name_cell.font = st["normal_font"]
        name_cell.fill = st["gray"]
        name_cell.border = st["border"]

        value_cell = ws.cell(row=row_idx, column=2, value=value)
        value_cell.font = st["bold_font"]
        value_cell.border = st["border"]
        value_cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 14

    _roster_sheet(
        wb, f"Registered - {label} Not Filled",
        f"{len(not_filled)} student(s) registered who have not filled the "
        f"{label.lower()} survey — {dept_name or 'All Departments'}",
        not_filled, st, st["gold"],
    )
    _roster_sheet(
        wb, f"Filled - {label} Survey",
        f"{len(filled)} student(s) who have filled the {label.lower()} survey — "
        f"{dept_name or 'All Departments'}",
        filled, st, st["navy"],
    )

    if dept_rows:
        _write_dept_breakdown(wb, dept_rows, st["navy"], st["gold"], st["gray"],
                              st["white_font"], st["title_font"], st["bold_font"],
                              st["normal_font"], st["border"])

    out_buf = io.BytesIO()
    wb.save(out_buf)
    return out_buf.getvalue()


def _write_dept_breakdown(wb, dept_rows, navy_fill, gold_fill, gray_fill,
                          white_font, title_font, bold_font, normal_font,
                          thin_border) -> None:
    """Second sheet: one line per department, plus a totals line.

    Mirrors the shared department directory, so the numbers on the page and
    the numbers in the workbook always agree.
    """
    ws = wb.create_sheet("Department Breakdown", 0)
    ws.views.sheetView[0].showGridLines = True

    ws["A1"] = "HACRI-E2 Department Breakdown"
    ws["A1"].font = title_font
    ws["A2"] = "Every department, side by side"
    ws["A2"].font = bold_font

    headers = [
        "Department", "Registered", "Baseline Filled", "Post Filled",
        "Baseline Pending", "Post Pending", "Reminders Sent", "Clicked Mail",
        "Filled After Mail", "Avg PRE Literacy", "Avg PRE Readiness",
        "Avg POST Literacy", "Avg POST Readiness",
    ]
    header_row = 4
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=text)
        cell.font = white_font
        cell.fill = gold_fill if text.startswith(("Post", "Avg POST")) else navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[header_row].height = 30

    totals = {k: 0 for k in ("registered", "pre_done", "post_done", "pre_pending",
                             "post_pending", "reminders_sent", "clicked", "completed_after")}
    row_idx = header_row
    for row_idx, d in enumerate(dept_rows, start=header_row + 1):
        for key in totals:
            totals[key] += d.get(key, 0) or 0
        values = [
            d.get("dept", ""),
            d.get("registered", 0), d.get("pre_done", 0), d.get("post_done", 0),
            d.get("pre_pending", 0), d.get("post_pending", 0),
            d.get("reminders_sent", 0), d.get("clicked", 0), d.get("completed_after", 0),
            d.get("avg_lit_pre"), d.get("avg_read_pre"),
            d.get("avg_lit_post"), d.get("avg_read_post"),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx,
                           value="—" if val is None else val)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "center")

    total_row = row_idx + 1
    total_values = [
        "ALL DEPARTMENTS",
        totals["registered"], totals["pre_done"], totals["post_done"],
        totals["pre_pending"], totals["post_pending"],
        totals["reminders_sent"], totals["clicked"], totals["completed_after"],
        "", "", "", "",
    ]
    for col_idx, val in enumerate(total_values, start=1):
        cell = ws.cell(row=total_row, column=col_idx, value=val)
        cell.font = bold_font
        cell.fill = gray_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "center")

    ws.column_dimensions["A"].width = 44
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def _safe_sheet_name(name: str, used: set[str]) -> str:
    """A legal, unique Excel sheet name derived from a department name.

    Excel forbids `\\ / * ? : [ ]` in sheet names and caps them at 31
    characters; several department names in this cohort are longer than
    that, and some pairs (UG/PG variants of the same department) are
    identical once truncated. `used` tracks names already claimed
    (lowercased) across the workbook so a collision gets a numeric suffix
    instead of silently overwriting the earlier sheet.
    """
    cleaned = re.sub(r'[\\/*?:\[\]]', "", name or "").strip() or "Department"
    base = cleaned[:31]
    candidate = base
    suffix_n = 2
    while candidate.lower() in used:
        suffix = f" ({suffix_n})"
        candidate = base[: 31 - len(suffix)] + suffix
        suffix_n += 1
    used.add(candidate.lower())
    return candidate


def _write_charts_sheet(wb, st, rows: list[dict], totals: dict) -> None:
    """"Charts" tab: cohort completion split (pie) and per-department
    volumes (bar), both driven by small backing tables on the same sheet
    so the charts stay editable in Excel rather than baked-in images.
    """
    ws = wb.create_sheet("Charts")
    ws["A1"] = "Deeksharambh 2026 — Charts"
    ws["A1"].font = st["title_font"]
    ws["A2"] = "Cohort completion split and per-department volumes"
    ws["A2"].font = st["normal_font"]

    # --- Pie: cohort completion split ---------------------------------
    registered = totals.get("registered", 0)
    pre_done = totals.get("pre_done", 0)
    post_done = totals.get("post_done", 0)
    pie_data = [
        ("Both surveys done", post_done),
        ("Baseline only", max(0, pre_done - post_done)),
        ("Not started baseline", max(0, registered - pre_done)),
    ]
    pie_header_row = 4
    ws.cell(row=pie_header_row, column=1, value="Status").font = st["white_font"]
    ws.cell(row=pie_header_row, column=1).fill = st["navy"]
    ws.cell(row=pie_header_row, column=2, value="Students").font = st["white_font"]
    ws.cell(row=pie_header_row, column=2).fill = st["navy"]
    for i, (label, value) in enumerate(pie_data, start=pie_header_row + 1):
        ws.cell(row=i, column=1, value=label).font = st["normal_font"]
        ws.cell(row=i, column=2, value=value).font = st["normal_font"]
    pie_last_row = pie_header_row + len(pie_data)

    pie = PieChart()
    pie.title = "Cohort completion split"
    pie_labels = Reference(ws, min_col=1, min_row=pie_header_row + 1, max_row=pie_last_row)
    pie_values = Reference(ws, min_col=2, min_row=pie_header_row, max_row=pie_last_row)
    pie.add_data(pie_values, titles_from_data=True)
    pie.set_categories(pie_labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.height, pie.width = 9, 14
    ws.add_chart(pie, "D4")

    # --- Bar: registered / baseline / post / Deeksharambh per department --
    bar_header_row = pie_last_row + 3
    bar_headers = ["Department", "Registered", "Baseline done", "Post survey done", "Deeksharambh done"]
    for col_idx, text in enumerate(bar_headers, start=1):
        cell = ws.cell(row=bar_header_row, column=col_idx, value=text)
        cell.font = st["white_font"]
        cell.fill = st["navy"]
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, row in enumerate(rows, start=bar_header_row + 1):
        ws.cell(row=i, column=1, value=row["dept"]).font = st["normal_font"]
        ws.cell(row=i, column=2, value=row.get("registered", 0)).font = st["normal_font"]
        ws.cell(row=i, column=3, value=row.get("pre_done", 0)).font = st["normal_font"]
        ws.cell(row=i, column=4, value=row.get("post_done", 0)).font = st["normal_font"]
        ws.cell(row=i, column=5, value=row.get("orientation_done", 0)).font = st["normal_font"]
    bar_last_row = bar_header_row + len(rows)

    if rows:
        bar = BarChart()
        bar.type = "bar"
        bar.title = "Departments by registered — descending"
        bar.y_axis.title = "Students"
        bar.x_axis.title = "Department"
        data = Reference(ws, min_col=2, max_col=5, min_row=bar_header_row, max_row=bar_last_row)
        cats = Reference(ws, min_col=1, min_row=bar_header_row + 1, max_row=bar_last_row)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        # openpyxl's x_axis is always the category axis (regardless of
        # bar vs col orientation) — reversing it puts the top row of the
        # backing table, the highest-registered department, at the top
        # of the chart instead of the bottom.
        bar.x_axis.scaling.orientation = "maxMin"
        bar.height = max(9, 1.1 * len(rows))
        bar.width = 22
        ws.add_chart(bar, f"D{pie_last_row + 3}")

    ws.column_dimensions["A"].width = 40
    for col in "BCDE":
        ws.column_dimensions[col].width = 16


def _write_department_sheet(wb, st, used_names: set[str], dept: str, row: dict, students: list[dict]) -> None:
    """One tab per department: its summary row, then every student behind it."""
    ws = wb.create_sheet(_safe_sheet_name(dept, used_names))

    ws["A1"] = dept
    ws["A1"].font = st["title_font"]
    ws["A2"] = (
        f"Registered {row.get('registered', 0)} · Baseline {row.get('pre_done', 0)} done, "
        f"{row.get('pre_pending', 0)} pending · Post survey {row.get('post_done', 0)} done, "
        f"{row.get('post_pending', 0)} pending · Deeksharambh {row.get('orientation_done', 0)} done, "
        f"{row.get('orientation_pending', 0)} pending"
    )
    ws["A2"].font = st["normal_font"]

    headers = ["Name", "Email", "Level", "Registered", "Baseline done", "Baseline date",
               "Post survey done", "Post survey date", "Deeksharambh done", "Deeksharambh date"]
    header_row = 4
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=text)
        cell.font = st["white_font"]
        cell.fill = st["navy"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = st["border"]
    ws.row_dimensions[header_row].height = 30

    def _fmt(when) -> str:
        try:
            return when.strftime("%d %b %Y, %H:%M")
        except AttributeError:
            return "—"

    for row_idx, s in enumerate(students, start=header_row + 1):
        values = [
            s.get("name") or "—",
            s.get("email") or "—",
            s.get("level") or "—",
            _fmt(s.get("registered_at")),
            "Yes" if s.get("pre_done") else "No", _fmt(s.get("pre_at")),
            "Yes" if s.get("post_done") else "No", _fmt(s.get("post_at")),
            "Yes" if s.get("orientation_done") else "No", _fmt(s.get("orientation_at")),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = st["normal_font"]
            cell.border = st["border"]
            cell.alignment = Alignment(horizontal="left" if col_idx in (1, 2) else "center")

    if not students:
        cell = ws.cell(row=header_row + 1, column=1, value="No students in this department")
        cell.font = st["normal_font"]

    widths = [26, 32, 8, 18, 12, 18, 14, 18, 16, 18]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = f"A{header_row}:J{header_row + len(students)}"


def generate_full_report_excel(report: dict, *, generated_at: str = "") -> bytes:
    """The "whole report in one Excel": overview, charts, then one tab per
    department (descending by registered count) with that department's
    summary and every student behind it.

    `report` is a `department_full_report()` result — `summary` (the same
    shape `generate_department_summary_excel` consumes) plus `departments`
    (each department's student list, already in the same descending order
    as the summary rows).
    """
    st = _styles()
    wb = Workbook()
    wb.remove(wb.active)

    summary = report.get("summary") or {}
    rows = summary.get("departments") or []
    totals = summary.get("totals") or {}

    overview = wb.create_sheet("Overview")
    _fill_department_summary_sheet(overview, st, summary, generated_at)

    _write_charts_sheet(wb, st, rows, totals)

    row_by_dept = {r["dept"]: r for r in rows}
    used_names: set[str] = {"overview", "charts"}
    for entry in report.get("departments") or []:
        dept = entry["dept"]
        _write_department_sheet(
            wb, st, used_names, dept,
            row_by_dept.get(dept, {}), entry.get("students") or [],
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
