"""
Per-user scorecard CSV.

Columns mirror the CLI's scorecard (Pre/Post Lit/Read/Quadrant/Band plus
deltas) and additionally include every Likert item in the SCHEMA so the
student can see exactly which items moved.
"""

from __future__ import annotations

import csv
import io
from typing import Any

from app.hacri_e2_compat import SCHEMA
from app.scoring import score_for_user

CSV_COLUMNS = [
    "Name",
    "Email",
    "PRE Literacy",
    "PRE Readiness",
    "PRE Quadrant",
    "PRE Band",
    "POST Literacy",
    "POST Readiness",
    "POST Quadrant",
    "POST Band",
    "Δ Literacy",
    "Δ Readiness",
    "Movement",
]


def _scorecard_row(name: str, email: str, pre: dict, post: dict) -> dict[str, Any]:
    pre_s = score_for_user(pre)
    post_s = score_for_user(post)
    d_lit = (
        round(post_s["lit"] - pre_s["lit"], 3)
        if (pre_s["lit"] is not None and post_s["lit"] is not None)
        else ""
    )
    d_read = (
        round(post_s["read"] - pre_s["read"], 3)
        if (pre_s["read"] is not None and post_s["read"] is not None)
        else ""
    )
    if isinstance(d_lit, float) and isinstance(d_read, float):
        if d_lit > 0 and d_read > 0:
            movement = "Champion gain (both up)"
        elif d_lit < 0 and d_read < 0:
            movement = "Decline (both down)"
        else:
            movement = "Mixed change"
    else:
        movement = "Insufficient data"

    return {
        "Name": name,
        "Email": email,
        "PRE Literacy": pre_s["lit"],
        "PRE Readiness": pre_s["read"],
        "PRE Quadrant": pre_s["quadrant"],
        "PRE Band": pre_s["band"],
        "POST Literacy": post_s["lit"],
        "POST Readiness": post_s["read"],
        "POST Quadrant": post_s["quadrant"],
        "POST Band": post_s["band"],
        "Δ Literacy": d_lit,
        "Δ Readiness": d_read,
        "Movement": movement,
    }


def _per_item_columns() -> list[str]:
    cols: list[str] = []
    for key in SCHEMA:
        cols += [f"PRE {key}", f"POST {key}", f"Δ {key}"]
    return cols


def scorecard_csv_bytes(name: str, email: str, pre: dict, post: dict) -> bytes:
    """UTF-8 CSV for one user."""
    row = _scorecard_row(name, email, pre, post)
    all_cols = CSV_COLUMNS + _per_item_columns()

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=all_cols, extrasaction="ignore")
    writer.writeheader()

    # Add per-item values
    for key in SCHEMA:
        pv = pre.get(key)
        ov = post.get(key)
        try:
            pv_n = int(pv) if pv not in (None, "") else ""
            ov_n = int(ov) if ov not in (None, "") else ""
            d = (ov_n - pv_n) if isinstance(pv_n, int) and isinstance(ov_n, int) else ""
        except (TypeError, ValueError):
            pv_n, ov_n, d = "", "", ""
        row[f"PRE {key}"] = pv_n
        row[f"POST {key}"] = ov_n
        row[f"Δ {key}"] = d

    writer.writerow(row)
    return buf.getvalue().encode("utf-8")


def cohort_csv_bytes(users_list: list[dict], pre_docs: list[dict], post_docs: list[dict]) -> bytes:
    """UTF-8 CSV for the entire cohort."""
    pre_map = {doc["email"]: doc.get("fields", {}) for doc in pre_docs}
    post_map = {doc["email"]: doc.get("fields", {}) for doc in post_docs}

    all_cols = CSV_COLUMNS + _per_item_columns()
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=all_cols, extrasaction="ignore")
    writer.writeheader()

    for u in users_list:
        email = u["email"]
        name = u.get("name", "")
        pre = pre_map.get(email, {})
        post = post_map.get(email, {})

        row = _scorecard_row(name, email, pre, post)

        # Add per-item values
        for key in SCHEMA:
            pv = pre.get(key)
            ov = post.get(key)
            try:
                pv_n = int(pv) if pv not in (None, "") else ""
                ov_n = int(ov) if ov not in (None, "") else ""
                d = (ov_n - pv_n) if isinstance(pv_n, int) and isinstance(ov_n, int) else ""
            except (TypeError, ValueError):
                pv_n, ov_n, d = "", "", ""
            row[f"PRE {key}"] = pv_n
            row[f"POST {key}"] = ov_n
            row[f"Δ {key}"] = d

        writer.writerow(row)

    return buf.getvalue().encode("utf-8")


def custom_cohort_export(
    users_list: list[dict],
    pre_docs: list[dict],
    post_docs: list[dict],
    format: str = "xlsx",
    inc_profile: bool = True,
    inc_timestamps: bool = True,
    inc_scores: bool = True,
    inc_responses: bool = False,
) -> tuple[bytes, str, str]:
    """Generates custom Excel or CSV files based on chosen filters and columns."""
    from datetime import datetime

    def _fmt_dt(dt: Any) -> str:
        if isinstance(dt, datetime):
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        elif isinstance(dt, str):
            return dt
        return ""

    pre_map = {doc["email"]: doc.get("fields", {}) for doc in pre_docs}
    post_map = {doc["email"]: doc.get("fields", {}) for doc in post_docs}

    # Build columns/headers list
    headers = []
    if inc_profile:
        headers += ["Name", "Email", "Level", "Programme", "Education Type"]
    if inc_timestamps:
        headers += ["Registered At", "Pre Submitted At", "Post Submitted At"]
    if inc_scores:
        headers += [
            "PRE Literacy", "PRE Readiness", "PRE Quadrant", "PRE Band",
            "POST Literacy", "POST Readiness", "POST Quadrant", "POST Band",
            "Δ Literacy", "Δ Readiness", "Movement"
        ]
    if inc_responses:
        for key in SCHEMA:
            headers += [f"PRE {key}", f"POST {key}", f"Δ {key}"]

    rows = []
    for u in users_list:
        email = u["email"]
        name = u.get("name", "")
        level = u.get("ug_or_pg", "ug").upper()
        edu = u.get("education_type", "")
        
        pre = pre_map.get(email, {})
        post = post_map.get(email, {})
        
        row = {}
        if inc_profile:
            row["Name"] = name
            row["Email"] = email
            row["Level"] = level
            row["Programme"] = u.get("program", "")
            row["Education Type"] = edu
            
        if inc_timestamps:
            row["Registered At"] = _fmt_dt(u.get("created_at"))
            row["Pre Submitted At"] = _fmt_dt(u.get("pre_submitted_at"))
            row["Post Submitted At"] = _fmt_dt(u.get("post_submitted_at"))
            
        if inc_scores:
            pre_s = score_for_user(pre)
            post_s = score_for_user(post)
            
            d_lit = (
                round(post_s["lit"] - pre_s["lit"], 3)
                if (pre_s["lit"] is not None and post_s["lit"] is not None)
                else ""
            )
            d_read = (
                round(post_s["read"] - pre_s["read"], 3)
                if (pre_s["read"] is not None and post_s["read"] is not None)
                else ""
            )
            
            if isinstance(d_lit, float) and isinstance(d_read, float):
                if d_lit > 0 and d_read > 0:
                    movement = "Champion gain"
                elif d_lit < 0 and d_read < 0:
                    movement = "Decline"
                else:
                    movement = "Mixed change"
            else:
                movement = "Insufficient data"
                
            row["PRE Literacy"] = pre_s["lit"] if pre_s["lit"] is not None else ""
            row["PRE Readiness"] = pre_s["read"] if pre_s["read"] is not None else ""
            row["PRE Quadrant"] = pre_s["quadrant"] if pre_s["quadrant"] is not None else ""
            row["PRE Band"] = pre_s["band"] if pre_s["band"] is not None else ""
            row["POST Literacy"] = post_s["lit"] if post_s["lit"] is not None else ""
            row["POST Readiness"] = post_s["read"] if post_s["read"] is not None else ""
            row["POST Quadrant"] = post_s["quadrant"] if post_s["quadrant"] is not None else ""
            row["POST Band"] = post_s["band"] if post_s["band"] is not None else ""
            row["Δ Literacy"] = d_lit
            row["Δ Readiness"] = d_read
            row["Movement"] = movement
            
        if inc_responses:
            for key in SCHEMA:
                pv = pre.get(key)
                ov = post.get(key)
                try:
                    pv_n = int(pv) if pv not in (None, "") else ""
                    ov_n = int(ov) if ov not in (None, "") else ""
                    d = (ov_n - pv_n) if isinstance(pv_n, int) and isinstance(ov_n, int) else ""
                except (TypeError, ValueError):
                    pv_n, ov_n, d = "", "", ""
                row[f"PRE {key}"] = pv_n
                row[f"POST {key}"] = ov_n
                row[f"Δ {key}"] = d
                
        rows.append(row)

    if format == "csv":
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
        return buf.getvalue().encode("utf-8"), "text/csv", "csv"
        
    else: # xlsx
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Cohort Export"
        ws.views.sheetView[0].showGridLines = True
        
        # Styles
        navy_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
        gold_fill = PatternFill(start_color="C9A84C", end_color="C9A84C", fill_type="solid")
        white_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        normal_font = Font(name="Segoe UI", size=11)
        
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )
        
        # Write Headers
        for col_idx, h_text in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h_text)
            cell.font = white_font
            if "POST" in h_text or "Δ" in h_text or "Movement" in h_text:
                cell.fill = gold_fill
            else:
                cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            
        ws.row_dimensions[1].height = 28
        
        # Write Data
        for row_idx, r_data in enumerate(rows, start=2):
            for col_idx, h_text in enumerate(headers, start=1):
                val = r_data.get(h_text, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = normal_font
                cell.border = thin_border
                
                # Alignments
                if h_text in ("Name", "Email", "Education Type", "Programme"):
                    cell.alignment = Alignment(horizontal="left")
                elif h_text in ("Level", "PRE Quadrant", "PRE Band", "POST Quadrant", "POST Band", "Movement"):
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="right")
                    
        # Auto-adjust widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
        out_buf = io.BytesIO()
        wb.save(out_buf)
        return out_buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"