"""
orientation_excel.py — the Deeksharambh orientation report as a workbook.

The same figures the deck prints and the shared page draws, in the form
somebody asked for when they wanted to sort a column: coverage first, then
every department, then the feedback questions with their denominators, then
every question option by option.

Built from `build_report()` / `department_rows()` output, so a number here can
never disagree with the same number on a slide. It carries no student rows —
this is the aggregate report, and the roster export lives in `excel_export`.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.orientation_analysis import MIN_REPORTABLE

NAVY = "1B2A4A"
GOLD = "C9A84C"
MIST = "F5F6FA"

HEAD_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
BAND_FILL = PatternFill(start_color=MIST, end_color=MIST, fill_type="solid")
HEAD_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Segoe UI", size=15, bold=True, color=NAVY)
NOTE_FONT = Font(name="Segoe UI", size=10, italic=True, color="5B6570")
BOLD = Font(name="Segoe UI", size=11, bold=True)
BODY = Font(name="Segoe UI", size=11)

WITHHELD = "withheld (n < %d)" % MIN_REPORTABLE


def _sheet(wb, title: str, heading: str, note: str = ""):
    ws = wb.create_sheet(title[:31])
    ws["A1"] = heading
    ws["A1"].font = TITLE_FONT
    if note:
        ws["A2"] = note
        ws["A2"].font = NOTE_FONT
    ws.freeze_panes = "A5"
    return ws


def _table(ws, row: int, headers: list[str], rows: list[list], widths: list[int]):
    for col, (name, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=col, value=name)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    for i, values in enumerate(rows, start=1):
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row + i, column=col, value=value)
            cell.font = BODY
            if i % 2 == 0:
                cell.fill = BAND_FILL
    return row + len(rows) + 2


def generate_orientation_excel(
    *, campus: str, scope: str, report: dict, departments: list[dict],
    campuses: list[dict] | None = None, generated_at: str = "",
) -> bytes:
    """The workbook behind the deck. `report` is a build_report() result."""
    head = report.get("headline", {})
    cover = report.get("coverage", {})
    highlights = report.get("highlights", {})
    answered_by = report.get("highlights_answered", {})
    labels = report.get("highlights_labels", {})
    total = report.get("count", 0)

    wb = Workbook()
    wb.remove(wb.active)

    # ── Overview ─────────────────────────────────────────────────────────────
    ws = _sheet(wb, "Overview", f"Deeksharambh 2026 — {campus} · {scope}",
                f"Generated {generated_at}." if generated_at else "")
    row = 5
    ws.cell(row=row, column=1, value="Coverage").font = BOLD
    row = _table(ws, row + 1, ["Measure", "Value"], [
        ["Responses", total],
        ["Students in scope", cover.get("eligible", 0)],
        ["Response rate", f"{cover.get('pct', 0):.1f}%"],
        ["Finished the baseline, not yet the orientation form", cover.get("pending", 0)],
        ["Departments answering", report.get("department_count", 0)],
        ["Replies with no matching student record", report.get("unmatched", 0)],
    ], [52, 18])

    ws.cell(row=row, column=1, value="Headline scores").font = BOLD
    row = _table(ws, row + 1, ["Measure", "Value", "Answered"], [
        ["Overall vibe (/10)", head.get("vibe"), _q(report, "q2")],
        ["Sense of belonging (/10)", head.get("belonging"), _q(report, "q29")],
        ["Confidence of succeeding (/10)", head.get("success"), _q(report, "q32")],
        ["Bridge Course confidence (/5)", head.get("bridge"), _q(report, "q16")],
        ["Recommendation score, mean (/10)", head.get("nps_avg"), head.get("nps_answered", 0)],
        ["Net Promoter Score", head.get("nps"), head.get("nps_answered", 0)],
        ["Promoters (9–10)", head.get("promoters", 0), head.get("nps_answered", 0)],
        ["Passives (7–8)", head.get("passives", 0), head.get("nps_answered", 0)],
        ["Detractors (0–6)", head.get("detractors", 0), head.get("nps_answered", 0)],
    ], [52, 18, 14])

    if campuses:
        ws.cell(row=row, column=1, value="By campus").font = BOLD
        _table(ws, row + 1, ["Campus", "Responses", "UG", "PG", "Departments"],
               [[c["campus"], c["filled"], c.get("ug", 0), c.get("pg", 0),
                 c.get("departments", 0)] for c in campuses], [24, 14, 10, 10, 16])

    # ── Departments ──────────────────────────────────────────────────────────
    ws = _sheet(wb, "Departments",
                "Every department with a student in scope",
                f"Scores are withheld where fewer than {MIN_REPORTABLE} students in the "
                "department answered: an average of one or two replies identifies the "
                "students who gave them. The counts are always shown.")
    ranked = sorted(departments,
                    key=lambda r: (r.get("vibe") is None, -(r.get("vibe") or 0),
                                   -r.get("filled", 0), r["dept"].lower()))
    _table(ws, 5,
           ["Department", "Answered", "In scope", "Pending", "Rate",
            "Vibe /10", "Belonging /10", "Succeeding /10", "Bridge /5", "NPS",
            "Promoters", "Passives", "Detractors", "Answered the NPS question"],
           [[r["dept"], r.get("filled", 0), r.get("eligible", 0), r.get("pending", 0),
             f"{r.get('pct', 0):.0f}%"] +
            [(WITHHELD if not r.get("reportable", True) else r.get(key))
             for key in ("vibe", "belonging", "success", "bridge", "nps",
                         "promoters", "passives", "detractors", "nps_answered")]
            for r in ranked],
           [46, 11, 11, 11, 9, 11, 15, 16, 11, 9, 12, 11, 12, 22])

    # ── Feedback ─────────────────────────────────────────────────────────────
    ws = _sheet(wb, "Feedback",
                "What students said to keep, stop, start and fix",
                "Every share is of the students who answered that one question — "
                "which is never the whole cohort. Both numbers are given.")
    rows = []
    for key, options in highlights.items():
        answered = answered_by.get(key, 0)
        for rank, option in enumerate(options or [], start=1):
            rows.append([labels.get(key, key), rank, option.get("label", ""),
                         option.get("count", 0), f"{option.get('pct', 0):.1f}%",
                         answered, total])
    _table(ws, 5,
           ["Question", "Rank", "Option", "Students", "Share of those who answered",
            "Answered this question", "Responses in total"],
           rows, [34, 8, 46, 12, 26, 22, 20])

    # ── Every question ───────────────────────────────────────────────────────
    ws = _sheet(wb, "All questions",
                "Every question, option by option",
                "Sections and questions in the order students saw them. "
                "Questions nobody answered are left out.")
    rows = []
    for section in report.get("sections", []):
        for question in section["questions"]:
            answered = question.get("answered", 0)
            if question.get("kind") == "matrix":
                for line in question.get("rows", []):
                    for option in line.get("options", []):
                        rows.append([section["title"], question["label"],
                                     line.get("label", ""), option.get("label", ""),
                                     option.get("count", 0),
                                     f"{option.get('pct', 0):.1f}%", answered])
                continue
            for option in question.get("options", []):
                rows.append([section["title"], question["label"], "",
                             option.get("label", ""), option.get("count", 0),
                             f"{option.get('pct', 0):.1f}%", answered])
    _table(ws, 5,
           ["Section", "Question", "Statement", "Option", "Students",
            "Share", "Answered this question"],
           rows, [26, 44, 34, 40, 12, 10, 22])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _q(report: dict, key: str) -> int:
    """How many students answered one question, by its key."""
    for section in report.get("sections", []):
        for question in section["questions"]:
            if question["key"] == key:
                return question.get("answered", 0)
    return 0
